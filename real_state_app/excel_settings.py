import numpy as np
import pandas as pd
import openpyxl as xl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side


def excel_save_setting(df, filename, isIndex=False, isHeader=False):

    try:
        df.to_excel(filename, index=isIndex, header=isHeader, encoding='utf-8-sig')
    except PermissionError as err:
        logger.error(err)
        excel_save = input('Excelを閉じた後、Y/nを選んでください。')
        if excel_save == 'Y':
            df.to_excel(filename, index=isIndex, header=isHeader, encoding='utf-8-sig')


def contain_index_header(dfHeader, dfContent):

    df = dfHeader.append(dfContent, ignore_index=True)
    index = [i+1 for i in range(len(dfContent))]
    index.insert(0, np.nan)
    df.insert(0, 'index', index)

    return df


def set_font(filename, font_name='メイリオ'):

    wb = xl.load_workbook(filename=filename)
    ws = wb.worksheets[0]
    font = Font(name=font_name)

    for row in ws:
        for cell in row:
            ws[cell.coordinate].font = font

    wb.save(filename)


def set_border(filename):

    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    wb = xl.load_workbook(filename=filename)
    ws = wb.worksheets[0]

    for row in ws:
        for cell in row:
            ws[cell.coordinate].border = border

    wb.save(filename)